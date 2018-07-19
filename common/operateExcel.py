# coding:utf-8
# @author : csl
# @date   : 2018/06/06 11:23
# excel表格操作封装

import openpyxl

class operateExcel(object):

    # 读取xlsx表格数据，返回列表
    def read_07_Excel(self, file_path, sheet_name):
        self.read_list = []
        self.rb = openpyxl.load_workbook(file_path)  # 打开文件
        self.sheet = self.rb[sheet_name]  # 通过sheet名称锁定表格
        for self.row in self.sheet.rows:  # 循环所有的行
            self.rowlist = []
            for self.cell in self.row:  # 循环行中所有的单元格
                # self.read_list.append(self.cell.value)
                self.rowlist.append(self.cell.value)
            self.read_list.append(self.rowlist)
        return self.read_list


    # 写入xlsx表格数据
    def write_07_Excel(self, file_path, value):
        self.wb = openpyxl.Workbook()  # 打开文件
        self.sheet = self.wb.active  # 激活sheet表格
        self.sheet.title = "测试结果"
        for self.i in range(0, len(value)):
            for self.j in range(0, len(value[self.i])):
                self.sheet.cell(row=self.i + 1, column=self.j + 1, value=str(value[self.i][self.j]))  # 写入单元格
        self.wb.save(file_path)
        print("写入07表格成功")

        pass


if __name__ == "__main__":
    # rw = operateExcel().read_07_Excel("../datas/userID.xlsx", "Sheet1")
    # print(rw)
    value = [["姓名", "年龄", "电话", "婚姻状况"],
             ["范彬彬", "22", "18888888888", "已婚"],
             ["袁姗姗", "25", "18999999999", "未婚"],
             ["刘德华", "50", "17777777777", "已婚"],
             ["张学友", "55", "15555555555", "已婚"],
             ["郭富城", "55", "13333333333", "已婚"]]
    wb = operateExcel().write_07_Excel("../datas/result.xlsx", value)